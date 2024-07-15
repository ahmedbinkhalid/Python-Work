#
# import requests
# from bs4 import BeautifulSoup
# import csv
# from urllib.parse import urljoin
#
# url = "https://mangaseek.net/person/list/a.html"
# max_tags = 20  # Maximum number of detail page URLs to process
# tag_count = 0
#
# # Open CSV file in write mode
# with open('works_diff(PN).csv', 'w', newline='', encoding='utf-8') as csvfile:
#     writer = csv.writer(csvfile)
#
#     response = requests.get(url)
#
#     if response.status_code == 200:
#         print("Successfully fetched the page content:")
#         soup = BeautifulSoup(response.text, 'html.parser')
#         rows = soup.find_all('tr')
#         for row in rows:
#             if tag_count >= max_tags:
#                 break
#             detail_page_url = "https://mangaseek.net" + row.find("a")["href"]
#             response = requests.get(detail_page_url)
#             if response.status_code == 200:
#                 page_soup = BeautifulSoup(response.text, 'html.parser')
#                 h3_tag = page_soup.find('h3', text='別ペンネームによる作品')
#                 if h3_tag:
#                     table = h3_tag.find_next('table', id='works')
#                     if table:
#                         # Extract table data
#                         rows = table.find_all('tr')
#                         for tr in rows:
#                             csv_row = [td.get_text(strip=True) for td in tr.find_all(['th', 'td'])]
#                             writer.writerow(csv_row)
#                         print(f"Data from 'works' table at {detail_page_url} written to CSV.")
#                     else:
#                         print(f"No 'works' table found at: {detail_page_url}")
#                     tag_count += 1
#                 else:
#                     print(f"No '別ペンネームによる作品' section found at: {detail_page_url}")
#             else:
#                 print(f"Failed to fetch page: {detail_page_url}. Status code:", response.status_code)
#     else:
#         print("Failed to fetch the page. Status code:", response.status_code)

# import requests
# from bs4 import BeautifulSoup
# import csv
# from urllib.parse import urljoin
#
# base_urls = ["https://mangaseek.net/person/list/a.html",
#              "https://mangaseek.net/person/list/ka.html",
#              "https://mangaseek.net/person/list/sa.html",
#              "https://mangaseek.net/person/list/ta.html",
#              "https://mangaseek.net/person/list/na.html",
#              "https://mangaseek.net/person/list/ha.html",
#              "https://mangaseek.net/person/list/ma.html",
#              "https://mangaseek.net/person/list/ya.html",
#              "https://mangaseek.net/person/list/ra.html",
#              "https://mangaseek.net/person/list/wa.html",
#              "https://mangaseek.net/person/list/tsu.html",
#              "https://mangaseek.net/person/list/me.html",
#              "https://mangaseek.net/person/list/i.html",
#              "https://mangaseek.net/person/list/no.html",
#              "https://mangaseek.net/person/list/ro.html"
#
#              ]
#
# max_rows = 20000  # Maximum number of rows to write to CSV
# row_count = 0
#
# # Open CSV file in write mode
# with open('works_diff(PN(All)).csv', 'w', newline='', encoding='utf-8') as csvfile:
#     writer = csv.writer(csvfile)
#
#     for url in base_urls:
#         response = requests.get(url)
#         if response.status_code == 200:
#             print("Successfully fetched the page content for URL:", url)
#             soup = BeautifulSoup(response.text, 'html.parser')
#             rows = soup.find_all('tr')
#             for row in rows:
#                 if row_count >= max_rows:
#                     break
#                 detail_page_url = "https://mangaseek.net" + row.find("a")["href"]
#                 response = requests.get(detail_page_url)
#                 if response.status_code == 200:
#                     page_soup = BeautifulSoup(response.text, 'html.parser')
#                     h3_tag = page_soup.find('h3', text='別ペンネームによる作品')
#                     if h3_tag:
#                         table = h3_tag.find_next('table', id='works')
#                         if table:
#                             # Extract table data
#                             table_rows = table.find_all('tr')
#                             for tr in table_rows:
#                                 csv_row = [td.get_text(strip=True) for td in tr.find_all(['th', 'td'])]
#                                 writer.writerow(csv_row)
#                                 print(f"Data from 'works' table at {detail_page_url} written to CSV.")
#                                 row_count += 1
#                                 if row_count >= max_rows:
#                                     break
#                         else:
#                             print(f"No 'works' table found at: {detail_page_url}")
#                     else:
#                         print(f"No '別ペンネームによる作品' section found at: {detail_page_url}")
#                 else:
#                     print(f"Failed to fetch page: {detail_page_url}. Status code:", response.status_code)
#             if row_count >= max_rows:
#                 break
#         else:
#             print(f"Failed to fetch the page for URL: {url}. Status code:", response.status_code)
#
# print("CSV writing complete.")

# import requests
# from bs4 import BeautifulSoup
# import csv
# from urllib.parse import urljoin
#
# base_urls = ["https://mangaseek.net/person/list/a.html",
#              "https://mangaseek.net/person/list/ka.html",
#              "https://mangaseek.net/person/list/sa.html",
#              "https://mangaseek.net/person/list/ta.html",
#              "https://mangaseek.net/person/list/na.html",
#              "https://mangaseek.net/person/list/ha.html",
#              "https://mangaseek.net/person/list/ma.html",
#              "https://mangaseek.net/person/list/ya.html",
#              "https://mangaseek.net/person/list/ra.html",
#              "https://mangaseek.net/person/list/wa.html",
#              "https://mangaseek.net/person/list/tsu.html",
#              "https://mangaseek.net/person/list/me.html",
#              "https://mangaseek.net/person/list/i.html",
#              "https://mangaseek.net/person/list/no.html",
#              "https://mangaseek.net/person/list/ro.html"]
#
# max_rows = 20000  # Maximum number of rows to write to CSV
# row_count = 0
#
# # Open CSV file in write mode
# with open('works_diff(PN).csv', 'w', newline='', encoding='utf-8') as csvfile:
#     writer = csv.writer(csvfile)
#
#     for url in base_urls:
#         response = requests.get(url)
#         if response.status_code == 200:
#             print("Successfully fetched the page content for URL:", url)
#             soup = BeautifulSoup(response.text, 'html.parser')
#             rows = soup.find_all('tr')
#             for row in rows:
#                 if row_count >= max_rows:
#                     break
#                 detail_page_url = "https://mangaseek.net" + row.find("a")["href"]
#                 response = requests.get(detail_page_url)
#                 if response.status_code == 200:
#                     page_soup = BeautifulSoup(response.text, 'html.parser')
#                     h3_tag = page_soup.find('h3', text='別ペンネームによる作品')
#                     if h3_tag:
#                         table = h3_tag.find_next('table', id='works')
#                         if table:
#                             print(f"Found 'works' table at: {detail_page_url}")
#                             # Extract table data
#                             table_rows = table.find_all('tr')
#                             for tr in table_rows:
#                                 csv_row = [td.get_text(strip=True) for td in tr.find_all(['th', 'td'])]
#                                 writer.writerow(csv_row)
#                                 row_count += 1
#                                 if row_count >= max_rows:
#                                     break
#                         else:
#                             print(f"No 'works' table found at: {detail_page_url}")
#                     else:
#                         print(f"No '別ペンネームによる作品' section found at: {detail_page_url}")
#                 else:
#                     print(f"Failed to fetch page: {detail_page_url}. Status code:", response.status_code)
#             if row_count >= max_rows:
#                 break
#         else:
#             print(f"Failed to fetch the page for URL: {url}. Status code:", response.status_code)
#
# print("CSV writing complete.")
# import requests
# from bs4 import BeautifulSoup
# import csv
# from urllib.parse import urljoin
# from time import sleep
#
# base_urls = [
#     "https://mangaseek.net/person/list/a.html",
#     # "https://mangaseek.net/person/list/ka.html",
#     # "https://mangaseek.net/person/list/sa.html",
#     # "https://mangaseek.net/person/list/ta.html",
#     # "https://mangaseek.net/person/list/na.html",
#     # "https://mangaseek.net/person/list/ha.html",
#     # "https://mangaseek.net/person/list/ma.html",
#     # "https://mangaseek.net/person/list/ya.html",
#     # "https://mangaseek.net/person/list/ra.html",
#     # "https://mangaseek.net/person/list/wa.html",
#     # "https://mangaseek.net/person/list/tsu.html",
#     # "https://mangaseek.net/person/list/me.html",
#     # "https://mangaseek.net/person/list/i.html",
#     # "https://mangaseek.net/person/list/no.html",
#     # "https://mangaseek.net/person/list/ro.html"
# ]
#
# max_rows = 20000  # Maximum number of rows to write to CSV
# row_count = 0
# retry_limit = 3  # Maximum number of retries for network requests
#
# # Open CSV file in write mode
# with open('works_diff(PN).xlsx', 'w', newline='', encoding='utf-8') as csvfile:
#     writer = csv.writer(csvfile)
#
#     for url in base_urls:
#         for attempt in range(retry_limit):
#             try:
#                 response = requests.get(url, timeout=10)
#                 response.raise_for_status()  # Raise HTTPError for bad responses
#                 break  # Exit retry loop if request is successful
#             except requests.exceptions.RequestException as e:
#                 print(f"Error fetching {url}: {e}")
#                 if attempt < retry_limit - 1:
#                     print(f"Retrying... ({attempt + 1}/{retry_limit})")
#                     sleep(5)  # Wait for 5 seconds before retrying
#                 else:
#                     print(f"Failed to fetch the page for URL: {url} after {retry_limit} attempts.")
#                     continue  # Skip to the next URL after max retries
#
#         if response.status_code == 200:
#             print("Successfully fetched the page content for URL:", url)
#             soup = BeautifulSoup(response.text, 'html.parser')
#             rows = soup.find_all('tr')
#             for row in rows:
#                 if row_count >= max_rows:
#                     break
#                 detail_page_url = "https://mangaseek.net" + row.find("a")["href"]
#                 for attempt in range(retry_limit):
#                     try:
#                         response = requests.get(detail_page_url, timeout=10)
#                         response.raise_for_status()
#                         break
#                     except requests.exceptions.RequestException as e:
#                         print(f"Error fetching {detail_page_url}: {e}")
#                         if attempt < retry_limit - 1:
#                             print(f"Retrying... ({attempt + 1}/{retry_limit})")
#                             sleep(5)
#                         else:
#                             print(f"Failed to fetch page: {detail_page_url} after {retry_limit} attempts.")
#                             continue
#
#                 if response.status_code == 200:
#                     page_soup = BeautifulSoup(response.text, 'html.parser')
#                     h3_tag = page_soup.find('h3', text='別ペンネームによる作品')
#                     if h3_tag:
#                         table = h3_tag.find_next('table', id='works')
#                         if table:
#                             print(f"Found 'works' table at: {detail_page_url}")
#                             # Extract table data
#                             table_rows = table.find_all('tr')
#                             for tr in table_rows:
#                                 csv_row = [td.get_text(strip=True) for td in tr.find_all(['th', 'td'])]
#                                 writer.writerow(csv_row)
#                                 row_count += 1
#                                 if row_count >= max_rows:
#                                     break
#                         else:
#                             print(f"No 'works' table found at: {detail_page_url}")
#                     else:
#                         print(f"No '別ペンネームによる作品' section found at: {detail_page_url}")
#                 else:
#                     print(f"Failed to fetch page: {detail_page_url}. Status code:", response.status_code)
#             if row_count >= max_rows:
#                 break
#         else:
#             print(f"Failed to fetch the page for URL: {url}. Status code:", response.status_code)
#
# print("CSV writing complete.")


# import requests
# from bs4 import BeautifulSoup
# from urllib.parse import urljoin
# from time import sleep
# from openpyxl import Workbook
#
# base_urls = [
#     "https://mangaseek.net/person/list/a.html",
#     # Add other URLs here if needed
# ]
#
# max_rows = 20000  # Maximum number of rows to write to Excel
# max_entries_per_page = 5  # Maximum number of entries to scrape from each detail page
# row_count = 0
# retry_limit = 3  # Maximum number of retries for network requests
#
# # Create a new Excel workbook and select the active worksheet
# workbook = Workbook()
# worksheet = workbook.active
# worksheet.title = "Works"
#
# def fetch_with_retries(url, retries=3, timeout=10):
#     for attempt in range(retries):
#         try:
#             response = requests.get(url, timeout=timeout)
#             response.raise_for_status()
#             return response
#         except requests.exceptions.RequestException as e:
#             print(f"Error fetching {url}: {e}")
#             if attempt < retries - 1:
#                 print(f"Retrying... ({attempt + 1}/{retries})")
#                 sleep(5)
#     return None
#
# for url in base_urls:
#     if row_count >= max_rows:
#         break
#
#     response = fetch_with_retries(url, retry_limit)
#     if not response:
#         print(f"Skipping URL due to repeated errors: {url}")
#         continue
#
#     print("Successfully fetched the page content for URL:", url)
#     soup = BeautifulSoup(response.text, 'html.parser')
#     rows = soup.find_all('tr')
#
#     for row in rows:
#         if row_count >= max_rows:
#             break
#
#         detail_page_url = urljoin("https://mangaseek.net", row.find("a")["href"])
#         detail_response = fetch_with_retries(detail_page_url, retry_limit)
#         if not detail_response:
#             print(f"Skipping detail page due to repeated errors: {detail_page_url}")
#             continue
#
#         page_soup = BeautifulSoup(detail_response.text, 'html.parser')
#         h3_tag = page_soup.find('h3', text='別ペンネームによる作品')
#         if not h3_tag:
#             print(f"No '別ペンネームによる作品' section found at: {detail_page_url}")
#             continue
#
#         table = h3_tag.find_next('table', id='works')
#         if not table:
#             print(f"No 'works' table found at: {detail_page_url}")
#             continue
#
#         print(f"Found 'works' table at: {detail_page_url}")
#         table_rows = table.find_all('tr')
#         entries_count = 0
#
#         for tr in table_rows:
#             if entries_count >= max_entries_per_page or row_count >= max_rows:
#                 break
#             row_data = [td.get_text(strip=True) for td in tr.find_all(['th', 'td'])]
#             worksheet.append(row_data)
#             row_count += 1
#             entries_count += 1
#
#         if row_count >= max_rows:
#             break
#
# # Save the workbook
# workbook.save('works_diff(PN).xlsx')
# print("Excel file writing complete.")
# import requests
# from bs4 import BeautifulSoup
# from openpyxl import Workbook
# from urllib.parse import urljoin
#
# url = "https://mangaseek.net/person/list/a.html"
# max_tags = 5  # Maximum number of detail page URLs to process
# tag_count = 0
#
# # Create a workbook and select the active worksheet
# wb = Workbook()
# ws = wb.active
# ws.title = "Works"
#
# response = requests.get(url)
#
# if response.status_code == 200:
#     print("Successfully fetched the page content:")
#     soup = BeautifulSoup(response.text, 'html.parser')
#     rows = soup.find_all('tr')
#     for row in rows:
#         if tag_count >= max_tags:
#             break
#         detail_page_url = "https://mangaseek.net" + row.find("a")["href"]
#         response = requests.get(detail_page_url)
#         if response.status_code == 200:
#             page_soup = BeautifulSoup(response.text, 'html.parser')
#             h3_tag = page_soup.find('h3', text='別ペンネームによる作品')
#             if h3_tag:
#                 table = h3_tag.find_next('table', id='works')
#                 if table:
#                     # Extract table data
#                     rows = table.find_all('tr')
#                     for tr in rows:
#                         excel_row = [td.get_text(strip=True) for td in tr.find_all(['th', 'td'])]
#                         ws.append(excel_row)
#                     print(f"Data from 'works' table at {detail_page_url} written to Excel.")
#                 else:
#                     print(f"No 'works' table found at: {detail_page_url}")
#                 tag_count += 1
#             else:
#                 print(f"No '別ペンネームによる作品' section found at: {detail_page_url}")
#         else:
#             print(f"Failed to fetch page: {detail_page_url}. Status code:", response.status_code)
# else:
#     print("Failed to fetch the page. Status code:", response.status_code)
#
# # Save the workbook to an Excel file
# wb.save('works_diff(PN).xlsx')
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from time import sleep
import openpyxl

base_urls = [
    "https://mangaseek.net/person/list/a.html",
    "https://mangaseek.net/person/list/ka.html",
    "https://mangaseek.net/person/list/sa.html",
    "https://mangaseek.net/person/list/ta.html",
    "https://mangaseek.net/person/list/na.html",
    "https://mangaseek.net/person/list/ha.html",
    "https://mangaseek.net/person/list/ma.html",
    "https://mangaseek.net/person/list/ya.html",
    "https://mangaseek.net/person/list/ra.html",
    "https://mangaseek.net/person/list/wa.html",
    "https://mangaseek.net/person/list/tsu.html",
    "https://mangaseek.net/person/list/me.html",
    "https://mangaseek.net/person/list/i.html",
    "https://mangaseek.net/person/list/no.html",
    "https://mangaseek.net/person/list/ro.html"
]

max_rows = 20000  # Maximum number of rows to write to Excel
row_count = 0
retry_limit = 3  # Maximum number of retries for network requests

# Create a new Excel workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Different pen names"

for url in base_urls:
    for attempt in range(retry_limit):
        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()  # Raise HTTPError for bad responses
            break  # Exit retry loop if request is successful
        except requests.exceptions.RequestException as e:
            print(f"Error fetching {url}: {e}")
            if attempt < retry_limit - 1:
                print(f"Retrying... ({attempt + 1}/{retry_limit})")
                sleep(5)  # Wait for 5 seconds before retrying
            else:
                print(f"Failed to fetch the page for URL: {url} after {retry_limit} attempts.")
                continue  # Skip to the next URL after max retries

    if response.status_code == 200:
        print("Successfully fetched the page content for URL:", url)
        soup = BeautifulSoup(response.text, 'html.parser')
        rows = soup.find_all('tr')
        for row in rows:
            if row_count >= max_rows:
                break
            detail_page_url = "https://mangaseek.net" + row.find("a")["href"]
            for attempt in range(retry_limit):
                try:
                    response = requests.get(detail_page_url, timeout=10)
                    response.raise_for_status()
                    break
                except requests.exceptions.RequestException as e:
                    print(f"Error fetching {detail_page_url}: {e}")
                    if attempt < retry_limit - 1:
                        print(f"Retrying... ({attempt + 1}/{retry_limit})")
                        sleep(5)
                    else:
                        print(f"Failed to fetch page: {detail_page_url} after {retry_limit} attempts.")
                        continue

            if response.status_code == 200:
                page_soup = BeautifulSoup(response.text, 'html.parser')
                h3_tag = page_soup.find('h3', text='別ペンネームによる作品')
                if h3_tag:
                    table = h3_tag.find_next('table', id='works')
                    if table:
                        print(f"Found 'works' table at: {detail_page_url}")
                        # Extract table data
                        table_rows = table.find_all('tr')
                        for tr in table_rows:
                            excel_row = [td.get_text(strip=True) for td in tr.find_all(['th', 'td'])]
                            sheet.append(excel_row)
                            row_count += 1
                            if row_count >= max_rows:
                                break
                    else:
                        print(f"No 'works' table found at: {detail_page_url}")
                else:
                    print(f"No '別ペンネームによる作品' section found at: {detail_page_url}")
            else:
                print(f"Failed to fetch page: {detail_page_url}. Status code:", response.status_code)
        if row_count >= max_rows:
            break
    else:
        print(f"Failed to fetch the page for URL: {url}. Status code:", response.status_code)

# Save the workbook
workbook.save('Different pen names(Japanese).xlsx')
print("Excel writing complete.")
