# import requests
# from bs4 import BeautifulSoup
# from openpyxl import Workbook
# import re
# import time
#
# # Function to scrape data from a URL
# def scrape_data_from_url(url):
#     # Initialize a list to store the scraped data
#     scraped_data = []
#
#     retries = 3
#     initial_response = None
#     for attempt in range(retries):
#         try:
#             # Send a request to the URL and get the HTML content
#             initial_response = requests.get(url, timeout=10)
#             initial_response.raise_for_status()  # Raise an exception for HTTP errors
#             break
#         except requests.exceptions.RequestException as e:
#             print(f"Error fetching {url}: {e}")
#             if attempt < retries - 1:
#                 print(f"Retrying... ({attempt + 1}/{retries})")
#                 time.sleep(2)
#             else:
#                 print(f"Failed to fetch {url} after {retries} attempts.")
#                 return []
#
#     if not initial_response:
#         print(f"Skipping URL due to failure: {url}")
#         return []
#
#     html_content = initial_response.content
#
#     # Parse the HTML content with BeautifulSoup
#     soup = BeautifulSoup(html_content, "html.parser")
#
#     # Find the table containing the work data
#     table = soup.find("table", {"id": "work"})
#
#     if not table:
#         print(f"No table found on URL: {url}")
#         return []
#
#     # Iterate over the table rows
#     for row in table.find_all("tr")[1:]:
#         # Extract the work ID and title from the row
#         work_id = row.find("span", {"class": "badge badge-work"}).text
#         title = row.find("a").text.strip()
#
#         # Get the URL of the work's detail page
#         detail_page_url = "https://mangaseek.net" + row.find("a")["href"]
#
#         # Retry mechanism and error handling for detail pages
#         detail_page_html = None
#         for attempt in range(retries):
#             try:
#                 # Send a request to the detail page URL and get the HTML content
#                 detail_page_response = requests.get(detail_page_url, timeout=10)
#                 detail_page_response.raise_for_status()  # Raise an exception for HTTP errors
#                 detail_page_html = detail_page_response.content
#                 break
#             except requests.exceptions.RequestException as e:
#                 print(f"Error fetching {detail_page_url}: {e}")
#                 if attempt < retries - 1:
#                     print(f"Retrying... ({attempt + 1}/{retries})")
#                     time.sleep(2)
#                 else:
#                     print(f"Failed to fetch {detail_page_url} after {retries} attempts. Skipping.")
#
#         if not detail_page_html:
#             continue
#
#         # Parse the detail page HTML content with BeautifulSoup
#         detail_page_soup = BeautifulSoup(detail_page_html, "html.parser")
#
#         # Find all tables on the detail page
#         detail_tables = detail_page_soup.find_all("table")
#
#         # Initialize variables to store scraped data
#         scraped_row = [work_id, title]
#
#         # Extract data from each table
#         for detail_table in detail_tables:
#             rows = detail_table.find_all("tr")
#             for row in rows:
#                 cells = row.find_all(["td", "th"])
#                 for cell in cells:
#                     scraped_row.append(cell.text.strip())
#
#         # Append the data to the list
#         scraped_data.append(scraped_row)
#
#         # Check if we have collected 2000 entries, if so, break the loop
#         if len(scraped_data) >= 20000:
#             break
#
#     return scraped_data
#
# # List of URLs to scrape
# urls = [
#     "https://mangaseek.net/work/list/a.html",
#     "https://mangaseek.net/work/list/ka.html",
#     "https://mangaseek.net/work/list/i.html",
#     "https://mangaseek.net/work/list/ki.html",
#     "https://mangaseek.net/work/list/u.html",
#     "https://mangaseek.net/work/list/ku.html",
#     "https://mangaseek.net/work/list/e.html",
#     "https://mangaseek.net/work/list/ke.html",
#     "https://mangaseek.net/work/list/o.html",
#     "https://mangaseek.net/work/list/ko.html",
#     "https://mangaseek.net/work/list/sa.html",
#     "https://mangaseek.net/work/list/hu.html",
#     # Add more URLs for other pages
# ]
#
# # Create a new Excel workbook and select the active sheet
# wb = Workbook()
# ws = wb.active
# ws.title = "Work List Data"
#
# # Write the header row
# header = ["Work ID", "Title"]
# if urls:
#     sample_data = scrape_data_from_url(urls[0])
#     if sample_data:
#         header += ["Data" + str(i) for i in range(1, len(sample_data[0]))]
# ws.append(header)
#
# # Iterate over each URL
# for url in urls:
#     print(f"Fetching URL: {url}")
#
#     # Scrape data from the URL
#     scraped_data = scrape_data_from_url(url)
#
#     # Write the scraped data to the Excel file
#     for data_row in scraped_data:
#         ws.append(data_row)
#
# # Save the workbook to a file
# wb.save("work_list(Japanese).xlsx")
#
# print("Data scraping completed. Results saved to work_data.xlsx")
#
# import requests
# from bs4 import BeautifulSoup
# from openpyxl import Workbook
# import time
#
# # Function to scrape data from a URL
# def scrape_data_from_url(url):
#     # Initialize a list to store the scraped data
#     scraped_data = []
#
#     retries = 3
#     initial_response = None
#     for attempt in range(retries):
#         try:
#             # Send a request to the URL and get the HTML content
#             initial_response = requests.get(url, timeout=10)
#             initial_response.raise_for_status()  # Raise an exception for HTTP errors
#             break
#         except requests.exceptions.RequestException as e:
#             print(f"Error fetching {url}: {e}")
#             if attempt < retries - 1:
#                 print(f"Retrying... ({attempt + 1}/{retries})")
#                 time.sleep(2)
#             else:
#                 print(f"Failed to fetch {url} after {retries} attempts.")
#                 return []
#
#     if not initial_response:
#         print(f"Skipping URL due to failure: {url}")
#         return []
#
#     html_content = initial_response.content
#
#     # Parse the HTML content with BeautifulSoup
#     soup = BeautifulSoup(html_content, "html.parser")
#
#     # Find the table containing the work data
#     table = soup.find("table", {"id": "work"})
#
#     if not table:
#         print(f"No table found on URL: {url}")
#         return []
#
#     # Iterate over the table rows
#     for row in table.find_all("tr")[1:]:
#         # Extract the work ID and title from the row
#         work_id = row.find("span", {"class": "badge badge-work"}).text
#         title = row.find("a").text.strip()
#
#         # Get the URL of the work's detail page
#         detail_page_url = "https://mangaseek.net" + row.find("a")["href"]
#
#         # Retry mechanism and error handling for detail pages
#         detail_page_html = None
#         for attempt in range(retries):
#             try:
#                 # Send a request to the detail page URL and get the HTML content
#                 detail_page_response = requests.get(detail_page_url, timeout=10)
#                 detail_page_response.raise_for_status()  # Raise an exception for HTTP errors
#                 detail_page_html = detail_page_response.content
#                 break
#             except requests.exceptions.RequestException as e:
#                 print(f"Error fetching {detail_page_url}: {e}")
#                 if attempt < retries - 1:
#                     print(f"Retrying... ({attempt + 1}/{retries})")
#                     time.sleep(2)
#                 else:
#                     print(f"Failed to fetch {detail_page_url} after {retries} attempts. Skipping.")
#
#         if not detail_page_html:
#             continue
#
#         # Parse the detail page HTML content with BeautifulSoup
#         detail_page_soup = BeautifulSoup(detail_page_html, "html.parser")
#
#         # Find all tables on the detail page
#         detail_tables = detail_page_soup.find_all("table")
#
#         # Initialize variables to store scraped data
#         scraped_row = [work_id, title]
#
#         # Extract data from each table
#         for detail_table in detail_tables:
#             rows = detail_table.find_all("tr")
#             for row in rows:
#                 cells = row.find_all(["td", "th"])
#                 for cell in cells:
#                     scraped_row.append(cell.text.strip())
#
#         # Append the data to the list
#         scraped_data.append(scraped_row)
#
#         # Check if we have collected 2000 entries, if so, break the loop
#         if len(scraped_data) >= 2:
#             break
#
#     return scraped_data
#
# # List of URLs to scrape
# urls = [
#     "https://mangaseek.net/work/list/a.html",
#     "https://mangaseek.net/work/list/ka.html",
#     "https://mangaseek.net/work/list/i.html",
#     "https://mangaseek.net/work/list/ki.html",
#     "https://mangaseek.net/work/list/u.html",
#     "https://mangaseek.net/work/list/ku.html",
#     "https://mangaseek.net/work/list/e.html",
#     "https://mangaseek.net/work/list/ke.html",
#     "https://mangaseek.net/work/list/o.html",
#     "https://mangaseek.net/work/list/ko.html",
#     "https://mangaseek.net/work/list/sa.html",
#     "https://mangaseek.net/work/list/hu.html",
#     # Add more URLs for other pages
# ]
#
# # Create a new Excel workbook and select the active sheet
# wb = Workbook()
# ws = wb.active
# ws.title = "Work List Data"
#
# # Flag to check if header is written
# header_written = False
#
# # Iterate over each URL
# for url in urls:
#     print(f"Fetching URL: {url}")
#
#     # Scrape data from the URL
#     scraped_data = scrape_data_from_url(url)
#
#     if not scraped_data:
#         continue
#
#     # Write the header row once
#     if not header_written:
#         header = ["Work ID", "Title"]
#         header += ["Data" + str(i) for i in range(1, len(scraped_data[0]))]
#         ws.append(header)
#         header_written = True
#
#     # Write the scraped data to the Excel file
#     for data_row in scraped_data:
#         ws.append(data_row)
#
# # Save the workbook to a file
# wb.save("work_list(Japanese).xlsx")
#
# print("Data scraping completed. Results saved to work_list(Japanese).xlsx")


import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import time

# Function to scrape data from a URL
def scrape_data_from_url(url, max_rows, total_rows):
    # Initialize a list to store the scraped data
    scraped_data = []

    retries = 3
    initial_response = None
    for attempt in range(retries):
        try:
            # Send a request to the URL and get the HTML content
            initial_response = requests.get(url, timeout=10)
            initial_response.raise_for_status()  # Raise an exception for HTTP errors
            break
        except requests.exceptions.RequestException as e:
            print(f"Error fetching {url}: {e}")
            if attempt < retries - 1:
                print(f"Retrying... ({attempt + 1}/{retries})")
                time.sleep(2)
            else:
                print(f"Failed to fetch {url} after {retries} attempts.")
                return []

    if not initial_response:
        print(f"Skipping URL due to failure: {url}")
        return []

    html_content = initial_response.content

    # Parse the HTML content with BeautifulSoup
    soup = BeautifulSoup(html_content, "html.parser")

    # Find the table containing the work data
    table = soup.find("table", {"id": "work"})

    if not table:
        print(f"No table found on URL: {url}")
        return []

    # Iterate over the table rows
    for row in table.find_all("tr")[1:]:
        if total_rows >= max_rows:
            break

        # Extract the work ID and title from the row
        work_id = row.find("span", {"class": "badge badge-work"}).text
        title = row.find("a").text.strip()

        # Get the URL of the work's detail page
        detail_page_url = "https://mangaseek.net" + row.find("a")["href"]

        # Retry mechanism and error handling for detail pages
        detail_page_html = None
        for attempt in range(retries):
            try:
                # Send a request to the detail page URL and get the HTML content
                detail_page_response = requests.get(detail_page_url, timeout=10)
                detail_page_response.raise_for_status()  # Raise an exception for HTTP errors
                detail_page_html = detail_page_response.content
                break
            except requests.exceptions.RequestException as e:
                print(f"Error fetching {detail_page_url}: {e}")
                if attempt < retries - 1:
                    print(f"Retrying... ({attempt + 1}/{retries})")
                    time.sleep(2)
                else:
                    print(f"Failed to fetch {detail_page_url} after {retries} attempts. Skipping.")

        if not detail_page_html:
            continue

        # Parse the detail page HTML content with BeautifulSoup
        detail_page_soup = BeautifulSoup(detail_page_html, "html.parser")

        # Find all tables on the detail page
        detail_tables = detail_page_soup.find_all("table")

        # Initialize variables to store scraped data
        scraped_row = [work_id, title]

        # Extract data from each table
        for detail_table in detail_tables:
            rows = detail_table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                for cell in cells:
                    scraped_row.append(cell.text.strip())

        # Append the data to the list
        scraped_data.append(scraped_row)
        total_rows += 1

    return scraped_data, total_rows

# List of URLs to scrape
urls = [
    "https://mangaseek.net/work/list/a.html",
    "https://mangaseek.net/work/list/ka.html",
    "https://mangaseek.net/work/list/i.html",
    "https://mangaseek.net/work/list/ki.html",
    "https://mangaseek.net/work/list/u.html",
    "https://mangaseek.net/work/list/ku.html",
    "https://mangaseek.net/work/list/e.html",
    "https://mangaseek.net/work/list/ke.html",
    "https://mangaseek.net/work/list/o.html",
    "https://mangaseek.net/work/list/ko.html",
    "https://mangaseek.net/work/list/sa.html",
    "https://mangaseek.net/work/list/hu.html",
    # Add more URLs for other pages
]

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Work List Data"

# Flag to check if header is written
header_written = False

# Limit on the total number of rows
max_rows = 22000
total_rows = 0

# Iterate over each URL
for url in urls:
    if total_rows >= max_rows:
        break

    print(f"Fetching URL: {url}")

    # Scrape data from the URL
    scraped_data, total_rows = scrape_data_from_url(url, max_rows, total_rows)

    if not scraped_data:
        continue

    # Write the header row once
    if not header_written:
        header = ["Work ID", "Title"]
        header += ["Data" + str(i) for i in range(1, len(scraped_data[0]))]
        ws.append(header)
        header_written = True

    # Write the scraped data to the Excel file
    for data_row in scraped_data:
        ws.append(data_row)

# Save the workbook to a file
wb.save("work_list(Japanese).xlsx")

print(f"Data scraping completed. Results saved to work_list(Japanese).xlsx with {total_rows} rows.")
